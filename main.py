import streamlit as st
import anthropic
import base64
import io
import time
import hashlib

# PDF
try:
    from PyPDF2 import PdfReader
    PDF_OK = True
except ImportError:
    PDF_OK = False

# Word
try:
    from docx import Document as DocxDocument
    DOCX_OK = True
except ImportError:
    DOCX_OK = False

# Excel
try:
    import openpyxl
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

# PIL
try:
    from PIL import Image
    PIL_OK = True
except ImportError:
    PIL_OK = False

# ==============================================================
# 페이지 설정
# ==============================================================
st.set_page_config(
    page_title="Claude AI 학습 도우미",
    page_icon="🤖",
    layout="wide",
)

# ==============================================================
# CSS
# ==============================================================
st.markdown("""
<style>
    .stApp {
        background: linear-gradient(135deg, #0f0c29, #302b63, #24243e);
    }
    .main-header {
        text-align: center;
        padding: 1.5rem 0 0.5rem 0;
    }
    .main-header h1 { color: #fff; font-size: 2.2rem; font-weight: 800; }
    .main-header p { color: #a0aec0; font-size: 1rem; }

    .glass-card {
        background: rgba(255,255,255,0.06);
        border: 1px solid rgba(255,255,255,0.12);
        border-radius: 16px;
        padding: 1.2rem 1.5rem;
        margin-bottom: 1rem;
        backdrop-filter: blur(10px);
    }
    .user-msg {
        background: rgba(99,102,241,0.15);
        border-left: 3px solid #6366f1;
        border-radius: 0 12px 12px 0;
        padding: 0.8rem 1.2rem;
        margin-bottom: 0.5rem;
        color: #c7d2fe;
        font-weight: 600;
    }
    .ai-msg {
        background: rgba(255,255,255,0.04);
        border-left: 3px solid #10b981;
        border-radius: 0 12px 12px 0;
        padding: 0.8rem 1.2rem;
        margin-bottom: 1.2rem;
        color: #e2e8f0;
        line-height: 1.7;
    }
    .metric-row { display: flex; gap: 1rem; flex-wrap: wrap; }
    .metric-box {
        flex: 1; min-width: 100px;
        background: rgba(255,255,255,0.04);
        border-radius: 12px; padding: 0.8rem; text-align: center;
    }
    .metric-box .label { color: #a0aec0; font-size: 0.75rem; }
    .metric-box .value { color: #fff; font-size: 1.3rem; font-weight: 700; }

    .login-card {
        max-width: 420px; margin: 4rem auto;
        background: rgba(255,255,255,0.08);
        border: 1px solid rgba(255,255,255,0.15);
        border-radius: 20px; padding: 2.5rem;
        text-align: center; backdrop-filter: blur(12px);
    }
    .login-card h2 { color: #fff; }
    .login-card p { color: #a0aec0; }

    .file-badge {
        display: inline-block;
        background: rgba(99,102,241,0.2);
        color: #c7d2fe;
        padding: 0.3rem 0.8rem;
        border-radius: 8px;
        font-size: 0.8rem;
        margin: 0.2rem;
    }
    .tuning-item {
        background: rgba(255,255,255,0.04);
        border: 1px solid rgba(255,255,255,0.08);
        border-radius: 10px;
        padding: 0.6rem 1rem;
        margin-bottom: 0.5rem;
        color: #e2e8f0;
    }

    section[data-testid="stSidebar"] {
        background: rgba(15, 12, 41, 0.95);
    }
    .stButton > button { border-radius: 10px; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# ==============================================================
# API 키
# ==============================================================
try:
    API_KEY = st.secrets["ANTHROPIC_API_KEY"]
except Exception:
    st.error("⚠️ `ANTHROPIC_API_KEY`가 Secrets에 설정되지 않았습니다.")
    st.stop()

# ==============================================================
# 비밀번호 (간단 로그인)
# ==============================================================
try:
    APP_PASSWORD = st.secrets["APP_PASSWORD"]
except Exception:
    APP_PASSWORD = None  # 비밀번호 미설정 시 로그인 없이 사용

# ==============================================================
# 모델 설정
# ==============================================================
MODEL_OPTIONS = {
    "Claude Sonnet 4 (빠르고 효율적)": "claude-sonnet-4-20250514",
    "Claude Opus 4 (최고 성능)": "claude-opus-4-20250514",
}
MODEL_PRICING = {
    "claude-sonnet-4-20250514": {"input": 3.0, "output": 15.0},
    "claude-opus-4-20250514": {"input": 15.0, "output": 75.0},
}

# ==============================================================
# 파일 처리 함수들
# ==============================================================
def format_size(size):
    if size < 1024: return f"{size} B"
    elif size < 1024**2: return f"{size/1024:.1f} KB"
    elif size < 1024**3: return f"{size/1024**2:.1f} MB"
    else: return f"{size/1024**3:.2f} GB"

def is_image(mime):
    return mime in ["image/jpeg", "image/png", "image/gif", "image/webp"]

def extract_pdf(data):
    if not PDF_OK: return "[PDF 라이브러리 없음]"
    try:
        reader = PdfReader(io.BytesIO(data))
        texts = []
        for i, page in enumerate(reader.pages):
            t = page.extract_text()
            if t: texts.append(f"--- 페이지 {i+1} ---\n{t}")
        return "\n\n".join(texts) if texts else "[텍스트 추출 불가]"
    except Exception as e:
        return f"[PDF 오류: {e}]"

def extract_docx(data):
    if not DOCX_OK: return "[Word 라이브러리 없음]"
    try:
        doc = DocxDocument(io.BytesIO(data))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        return f"[Word 오류: {e}]"

def extract_xlsx(data):
    if not XLSX_OK: return "[Excel 라이브러리 없음]"
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        lines = []
        for name in wb.sheetnames:
            lines.append(f"=== 시트: {name} ===")
            for row in wb[name].iter_rows(values_only=True):
                r = "\t".join(str(c) if c else "" for c in row)
                if r.strip(): lines.append(r)
        wb.close()
        return "\n".join(lines)
    except Exception as e:
        return f"[Excel 오류: {e}]"

def process_file(uploaded):
    data = uploaded.read()
    mime = uploaded.type or ""
    name = uploaded.name
    size_str = format_size(len(data))

    if is_image(mime):
        b64 = base64.standard_b64encode(data).decode("utf-8")
        block = {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}}
        return block, f"[이미지: {name}]", f"🖼️ {name} ({size_str})"

    if mime == "application/pdf":
        text = extract_pdf(data)
    elif "wordprocessingml" in mime:
        text = extract_docx(data)
    elif "spreadsheetml" in mime:
        text = extract_xlsx(data)
    else:
        try: text = data.decode("utf-8")
        except:
            try: text = data.decode("euc-kr")
            except: text = f"[읽기 불가: {name}]"

    return None, text, f"📄 {name} ({size_str})"

# ==============================================================
# 세션 초기화
# ==============================================================
defaults = {
    "logged_in": False,
    "username": "",
    "messages": [],
    "total_input": 0,
    "total_output": 0,
    "total_cost": 0.0,
    "tuning_profile": {},
    "tuning_preferences": {},
    "tuning_knowledge": [],
    "tuning_instructions": [],
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ==============================================================
# 시스템 프롬프트 생성 (튜닝 데이터 포함)
# ==============================================================
def build_system_prompt():
    base = (
        "너는 당곡고등학교 학생들의 학습을 돕는 친절한 AI 도우미야. "
        "학생이 스스로 생각하고 탐구할 수 있도록 도와줘. "
        "설명은 고등학생 눈높이에 맞춰 쉽고 친근하게 해줘. "
        "한국어로 답변해줘.\n\n"
    )
    parts = [base]

    # 프로필
    profile = st.session_state.tuning_profile
    if profile:
        items = [f"- {k}: {v}" for k, v in profile.items() if v]
        if items:
            parts.append("【이 학생의 프로필】\n" + "\n".join(items))

    # 선호 설정
    prefs = st.session_state.tuning_preferences
    if prefs:
        items = [f"- {k}: {v}" for k, v in prefs.items() if v]
        if items:
            parts.append("【학습 선호 설정】\n" + "\n".join(items))

    # 배경 지식
    knowledge = st.session_state.tuning_knowledge
    if knowledge:
        items = [f"- {k['title']}: {k['content']}" for k in knowledge]
        parts.append("【배경 지식 / 메모】\n" + "\n".join(items))

    # 특별 지시
    instructions = st.session_state.tuning_instructions
    if instructions:
        parts.append("【특별 지시사항】\n" + "\n".join(f"- {i}" for i in instructions))

    return "\n\n".join(parts)

# ==============================================================
# 로그인 화면
# ==============================================================
if not st.session_state.logged_in:
    if APP_PASSWORD is None:
        # 비밀번호 미설정 → 바로 입장
        st.session_state.logged_in = True
        st.session_state.username = "사용자"
        st.rerun()

    st.markdown("""
    <div class="main-header">
        <h1>🤖 Claude AI 학습 도우미</h1>
        <p>당곡고등학교 전용 AI 학습 플랫폼</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="login-card"><h2>🔐 로그인</h2><p>이름과 비밀번호를 입력하세요</p></div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        input_name = st.text_input("이름", placeholder="예: 홍길동")
        input_pw = st.text_input("비밀번호", type="password", placeholder="비밀번호 입력")

        if st.button("🚀 로그인", use_container_width=True):
            if not input_name.strip():
                st.warning("이름을 입력해주세요!")
            elif input_pw == APP_PASSWORD:
                st.session_state.logged_in = True
                st.session_state.username = input_name.strip()
                st.rerun()
            else:
                st.error("❌ 비밀번호가 틀렸습니다!")

    st.stop()

# ==============================================================
# 메인 앱 (로그인 후)
# ==============================================================
username = st.session_state.username

# --- 사이드바 ---
with st.sidebar:
    st.markdown(f"### 👤 {username}님")

    if st.button("🚪 로그아웃", use_container_width=True):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

    st.markdown("---")

    # 모델 선택
    st.markdown("### 🧠 AI 모델")
    sel_label = st.selectbox("모델", list(MODEL_OPTIONS.keys()), index=0, label_visibility="collapsed")
    sel_model = MODEL_OPTIONS[sel_label]
    pricing = MODEL_PRICING[sel_model]
    st.caption(f"입력 ${pricing['input']}/1M · 출력 ${pricing['output']}/1M")

    st.markdown("---")

    # 사용량
    st.markdown("### 📊 사용량")
    st.markdown(f"- 입력: **{st.session_state.total_input:,}** 토큰")
    st.markdown(f"- 출력: **{st.session_state.total_output:,}** 토큰")
    st.markdown(f"- 비용: **${st.session_state.total_cost:.4f}**")

    st.markdown("---")

    if st.button("🗑️ 대화 초기화", use_container_width=True):
        st.session_state.messages = []
        st.session_state.total_input = 0
        st.session_state.total_output = 0
        st.session_state.total_cost = 0.0
        st.rerun()

    st.markdown("---")

    # 메뉴
    st.markdown("### 📌 메뉴")
    page = st.radio("페이지", ["💬 채팅", "🎯 AI 튜닝"], index=0, label_visibility="collapsed")

# --- 헤더 ---
st.markdown(f"""
<div class="main-header">
    <h1>🤖 Claude AI 학습 도우미</h1>
    <p>{username}님 환영합니다! | 모델: {sel_model}</p>
</div>
""", unsafe_allow_html=True)

# ==============================================================
# 채팅 페이지
# ==============================================================
if page == "💬 채팅":

    # 대화 기록 표시
    if st.session_state.messages:
        st.markdown("### 💬 대화 기록")
        for msg in st.session_state.messages:
            if msg["role"] == "user":
                display = msg.get("display", msg["content"])
                st.markdown(f'<div class="user-msg">🙋 {display[:500]}</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="ai-msg">{msg["content"]}</div>', unsafe_allow_html=True)
        st.markdown("---")

    # 파일 업로드
    st.markdown("### 📎 파일 첨부 (선택)")
    files = st.file_uploader(
        "이미지, PDF, Word, Excel, 텍스트 등",
        type=["jpg","jpeg","png","gif","webp",
              "pdf","docx","xlsx","txt","csv","md","json",
              "py","js","html","css","java","c","cpp"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    if files:
        for f in files:
            st.markdown(f'<span class="file-badge">📎 {f.name} ({format_size(f.size)})</span>', unsafe_allow_html=True)
            if f.type and is_image(f.type):
                st.image(f, width=300)

    # 질문 입력
    st.markdown("### ✏️ 질문")
    question = st.text_area("질문", placeholder="질문을 입력하세요...", height=120, label_visibility="collapsed")

    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        send = st.button("🚀 질문하기", use_container_width=True)

    if send and (question.strip() or files):
        with st.spinner("🤔 AI가 생각하고 있어요..."):
            try:
                # 파일 처리
                file_results = []
                if files:
                    for f in files:
                        f.seek(0)
                        file_results.append(process_file(f))

                # Claude 메시지 구성
                content_blocks = []

                # 이미지 블록
                for block, _, _ in file_results:
                    if block: content_blocks.append(block)

                # 텍스트 조합
                text_parts = []
                for block, extracted, _ in file_results:
                    if block is None and extracted:
                        text_parts.append(f"[첨부 파일 내용]\n{extracted}")
                if question.strip():
                    text_parts.append(question.strip())

                combined = "\n\n".join(text_parts) if text_parts else "첨부 파일을 분석해주세요."
                content_blocks.append({"type": "text", "text": combined})

                # 표시용 텍스트
                display_text = question.strip()
                if file_results:
                    names = [r[2] for r in file_results]
                    display_text = f"[📎 {', '.join(names)}]\n{display_text}"

                # 메시지 기록에 추가
                st.session_state.messages.append({
                    "role": "user",
                    "content": combined,
                    "display": display_text,
                })

                # API 호출용 메시지 구성
                api_msgs = []
                for i, msg in enumerate(st.session_state.messages):
                    if i == len(st.session_state.messages) - 1 and msg["role"] == "user":
                        api_msgs.append({"role": "user", "content": content_blocks})
                    else:
                        api_msgs.append({"role": msg["role"], "content": msg["content"]})

                # API 호출
                client = anthropic.Anthropic(api_key=API_KEY)
                start = time.time()

                response = client.messages.create(
                    model=sel_model,
                    max_tokens=8192,
                    system=build_system_prompt(),
                    messages=api_msgs,
                )

                elapsed = time.time() - start
                answer = response.content[0].text
                inp_tok = response.usage.input_tokens
                out_tok = response.usage.output_tokens
                cost = (inp_tok * pricing["input"] / 1e6) + (out_tok * pricing["output"] / 1e6)

                # 저장
                st.session_state.messages.append({"role": "assistant", "content": answer})
                st.session_state.total_input += inp_tok
                st.session_state.total_output += out_tok
                st.session_state.total_cost += cost

                # 답변 표시
                st.markdown("### 💡 AI 답변")
                st.markdown(answer)

                # 사용량
                st.markdown(f"""
                <div class="glass-card">
                    <div style="color:#a0aec0;font-size:0.8rem;font-weight:600;margin-bottom:0.5rem;">📊 사용량</div>
                    <div class="metric-row">
                        <div class="metric-box">
                            <div class="label">입력</div>
                            <div class="value">{inp_tok:,}</div>
                        </div>
                        <div class="metric-box">
                            <div class="label">출력</div>
                            <div class="value">{out_tok:,}</div>
                        </div>
                        <div class="metric-box">
                            <div class="label">시간</div>
                            <div class="value">{elapsed:.1f}s</div>
                        </div>
                        <div class="metric-box">
                            <div class="label">비용</div>
                            <div class="value">${cost:.4f}</div>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            except anthropic.AuthenticationError:
                st.error("❌ API 키가 올바르지 않습니다.")
            except anthropic.RateLimitError:
                st.error("⏳ API 한도 초과. 잠시 후 다시 시도하세요.")
            except anthropic.APIError as e:
                st.error(f"❌ API 오류: {e}")
            except Exception as e:
                st.error(f"❌ 오류: {e}")
                import traceback
                st.code(traceback.format_exc())

    elif send:
        st.warning("⚠️ 질문을 입력하거나 파일을 첨부해주세요!")

# ==============================================================
# AI 튜닝 페이지
# ==============================================================
elif page == "🎯 AI 튜닝":
    st.markdown("### 🎯 나만의 AI 튜닝")
    st.markdown("""
    <div class="glass-card">
        <p style="color:#e2e8f0;">
            여기서 설정한 정보를 바탕으로 AI가 <b>나에게 맞춤 답변</b>을 해줍니다.<br>
            설정은 현재 세션(탭)에서 유지됩니다.
        </p>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs(["👤 프로필", "📚 학습 선호", "🧠 배경 지식", "✏️ 특별 지시"])

    # --- 프로필 ---
    with tab1:
        st.markdown("**나의 학습 프로필**")
        fields = {
            "학년": "예: 2학년",
            "관심 과목": "예: 수학, 물리",
            "진로 희망": "예: 컴퓨터공학과",
            "학습 수준": "예: 수학 상위권",
            "약점 과목": "예: 국어 비문학",
        }
        for field, ph in fields.items():
            val = st.text_input(
                field,
                value=st.session_state.tuning_profile.get(field, ""),
                placeholder=ph,
                key=f"prof_{field}",
            )
            st.session_state.tuning_profile[field] = val

    # --- 학습 선호 ---
    with tab2:
        st.markdown("**AI 답변 스타일 설정**")
        pref_opts = {
            "설명 방식": ["쉽고 친근하게", "학술적으로 정확하게", "예시 위주로", "단계별로 상세하게"],
            "답변 길이": ["짧고 핵심만", "적당한 길이", "길고 상세하게"],
            "언어 스타일": ["반말 (친근)", "존댓말 (정중)", "이모지 많이"],
            "수학 표현": ["텍스트로", "LaTeX 수식으로"],
        }
        for key, opts in pref_opts.items():
            cur = st.session_state.tuning_preferences.get(key, opts[0])
            idx = opts.index(cur) if cur in opts else 0
            sel = st.selectbox(key, opts, index=idx, key=f"pref_{key}")
            st.session_state.tuning_preferences[key] = sel

    # --- 배경 지식 ---
    with tab3:
        st.markdown("**AI가 기억할 배경 지식**")

        for i, item in enumerate(st.session_state.tuning_knowledge):
            col_a, col_b = st.columns([5, 1])
            with col_a:
                st.markdown(f'<div class="tuning-item">📌 <b>{item["title"]}</b>: {item["content"]}</div>', unsafe_allow_html=True)
            with col_b:
                if st.button("❌", key=f"del_k_{i}"):
                    st.session_state.tuning_knowledge.pop(i)
                    st.rerun()

        st.markdown("---")
        k_title = st.text_input("제목", placeholder="예: 수학 학습 현황", key="new_k_title")
        k_content = st.text_area("내용", placeholder="예: 미적분 단원 진행 중", key="new_k_content", height=80)
        if st.button("💾 추가", key="add_k"):
            if k_title.strip() and k_content.strip():
                st.session_state.tuning_knowledge.append({"title": k_title.strip(), "content": k_content.strip()})
                st.rerun()
            else:
                st.warning("제목과 내용을 모두 입력해주세요.")

    # --- 특별 지시 ---
    with tab4:
        st.markdown("**AI에게 내릴 특별 지시사항**")

        for i, inst in enumerate(st.session_state.tuning_instructions):
            col_a, col_b = st.columns([5, 1])
            with col_a:
                st.markdown(f'<div class="tuning-item">⚡ {inst}</div>', unsafe_allow_html=True)
            with col_b:
                if st.button("❌", key=f"del_i_{i}"):
                    st.session_state.tuning_instructions.pop(i)
                    st.rerun()

        st.markdown("---")
        new_inst = st.text_area("지시사항", placeholder="예: 항상 관련 공식을 먼저 정리해줘", key="new_inst", height=80)
        if st.button("💾 추가", key="add_i"):
            if new_inst.strip():
                st.session_state.tuning_instructions.append(new_inst.strip())
                st.rerun()
            else:
                st.warning("내용을 입력해주세요.")

    # 프롬프트 미리보기
    st.markdown("---")
    with st.expander("🔍 현재 시스템 프롬프트 미리보기"):
        st.code(build_system_prompt(), language="text")

# ==============================================================
# 하단
# ==============================================================
st.markdown("---")
st.markdown("""
<div style="text-align:center;color:#718096;font-size:0.85rem;padding:1rem 0;">
    🏫 당곡고등학교 AI 학습 도우미 | Claude API 기반
</div>
""", unsafe_allow_html=True)
